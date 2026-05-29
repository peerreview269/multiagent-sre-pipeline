"""
network_to_excel.py
===================
Converts a social network extraction JSON output (from run.py) into:
  - A formatted multisheet Excel report
  - UCINET DL (.txt) multiplex matrix file
  - iGraph-ready CSVs (edges + nodes)
  - NetworkX exports: edges CSV, nodes CSV, GraphML, GEXF, per-layer GraphML files

All outputs go into a dedicated folder named after the story.

Usage:
    python network_to_excel.py outputs/bartleby_result.json
    python network_to_excel.py outputs/bartleby_result.json --outdir reports/

Example:
    python network_to_excel.py outputs/bartleby_result.json --outdir reports/bartleby
"""

import argparse
import csv
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

import networkx as nx
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Font, PatternFill, Border, Side
)
from openpyxl.utils import get_column_letter

# ── Colour palette ────────────────────────────────────────────────────────────
DARK_BLUE   = "1F3864"
MID_BLUE    = "2E75B6"
LIGHT_BLUE  = "BDD7EE"
LIGHT_GREY  = "F2F2F2"
WHITE       = "FFFFFF"

TIE_COLORS = {
    "family":       "70AD47",   # green
    "friendship":   "00B0F0",   # cyan
    "romantic":     "FF69B4",   # pink
    "professional": "FFC000",   # amber
    "adversarial":  "FF0000",   # red
}

# ── Style helpers ─────────────────────────────────────────────────────────────
def hdr(cell, text, bold=True, color=WHITE, bg=DARK_BLUE, size=11, wrap=False, center=False):
    cell.value = text
    cell.font = Font(bold=bold, color=color, size=size, name="Arial")
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center",
        wrap_text=wrap,
    )

def sub_hdr(cell, text):
    hdr(cell, text, bg=MID_BLUE, size=10)

def body(cell, value, bold=False, bg=WHITE, wrap=False, center=False):
    cell.value = value
    cell.font = Font(name="Arial", size=10, bold=bold)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center",
        wrap_text=wrap,
    )

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_border(ws, min_row, max_row, min_col, max_col):
    b = thin_border()
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                             min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = b

def set_col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

def freeze(ws, cell="A2"):
    ws.freeze_panes = cell

def alt_row_bg(row_idx):
    return LIGHT_GREY if row_idx % 2 == 0 else WHITE

# ── Sheet builders ────────────────────────────────────────────────────────────

def build_summary(wb: Workbook, data: dict):
    ws = wb.create_sheet("Summary")
    summary = data.get("summary", {})
    story_name = data.get("story_name", "Unknown")

    # Title
    ws.merge_cells("A1:D1")
    hdr(ws["A1"], f"Social Network Report — {story_name}",
        size=14, center=True)
    ws.row_dimensions[1].height = 28

    # Key metrics table
    metrics = [
        ("Total Characters",    summary.get("total_characters", 0)),
        ("Total Relationships", summary.get("total_relationships", 0)),
        ("Isolates",            summary.get("isolate_count", 0)),
        ("Network Hub",         summary.get("network_hub", "—")),
        ("Critic Rounds",       summary.get("critique_rounds", 0)),
        ("ER Actions",          summary.get("entity_resolution_actions", 0)),
        ("Conflict Groups Reviewed", summary.get("conflict_groups_reviewed", 0)),
    ]
    ws.merge_cells("A3:D3")
    sub_hdr(ws["A3"], "Key Metrics")

    for i, (label, value) in enumerate(metrics, start=4):
        bg = alt_row_bg(i)
        body(ws[f"A{i}"], label, bold=True, bg=bg)
        body(ws[f"B{i}"], value, bg=bg)
        ws.merge_cells(f"B{i}:D{i}")

    apply_border(ws, 4, 3 + len(metrics), 1, 4)

    # Tie breakdown
    row = 4 + len(metrics) + 2
    ws.merge_cells(f"A{row}:D{row}")
    sub_hdr(ws[f"A{row}"], "Tie Breakdown")
    row += 1

    breakdown = summary.get("tie_breakdown", {})
    for tie_type in ["family", "friendship", "romantic", "professional", "adversarial"]:
        count = breakdown.get(tie_type, 0)
        bg = alt_row_bg(row)
        body(ws[f"A{row}"], tie_type.capitalize(), bg=bg)
        body(ws[f"B{row}"], count, bg=bg, center=True)
        # Colour swatch
        swatch_color = TIE_COLORS.get(tie_type, "CCCCCC")
        ws[f"C{row}"].fill = PatternFill("solid", start_color=swatch_color)
        ws.merge_cells(f"B{row}:D{row}")
        row += 1

    apply_border(ws, row - 5, row - 1, 1, 4)

    # Analyst notes
    notes = data.get("analyst_notes", [])
    if notes:
        row += 1
        ws.merge_cells(f"A{row}:D{row}")
        sub_hdr(ws[f"A{row}"], "Analyst Notes")
        row += 1
        for note in notes:
            ws.merge_cells(f"A{row}:D{row}")
            body(ws[f"A{row}"], f"• {note}", bg=alt_row_bg(row), wrap=True)
            ws.row_dimensions[row].height = 30
            row += 1
        apply_border(ws, row - len(notes), row - 1, 1, 4)

    set_col_widths(ws, {"A": 28, "B": 18, "C": 12, "D": 12})
    freeze(ws, "A2")


def build_characters(wb: Workbook, data: dict):
    ws = wb.create_sheet("Characters")
    characters = data.get("characters", [])
    isolate_ids = {c.get("id") for c in data.get("isolates", [])}

    headers = ["ID", "Name", "Aliases", "Role", "Type", "Isolate?"]
    for col, h in enumerate(headers, 1):
        sub_hdr(ws.cell(row=1, column=col), h)
    ws.row_dimensions[1].height = 18

    for i, char in enumerate(characters, start=2):
        bg = alt_row_bg(i)
        cid = char.get("id", "")
        is_iso = "Yes" if cid in isolate_ids else ""
        row_data = [
            cid,
            char.get("name", ""),
            ", ".join(char.get("aliases", [])),
            char.get("role", ""),
            char.get("type", "PERSON"),
            is_iso,
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col)
            body(cell, val, bg=bg, wrap=(col == 4))
            if col == 6 and is_iso:
                cell.font = Font(name="Arial", size=10, bold=True, color="FF0000")
        ws.row_dimensions[i].height = 30 if char.get("role") else 18

    apply_border(ws, 1, 1 + len(characters), 1, len(headers))
    set_col_widths(ws, {"A": 22, "B": 22, "C": 24, "D": 48, "E": 10, "F": 10})
    freeze(ws, "A2")
    ws.auto_filter.ref = f"A1:F{1 + len(characters)}"


def build_relationships(wb: Workbook, data: dict):
    ws = wb.create_sheet("Relationships")
    rels = data.get("relationships", [])

    headers = ["Source", "Target", "Tie Type", "Sub-Type", "Direction",
               "Strength", "Confidence", "Chunks", "Evidence"]
    for col, h in enumerate(headers, 1):
        sub_hdr(ws.cell(row=1, column=col), h)
    ws.row_dimensions[1].height = 18

    for i, rel in enumerate(rels, start=2):
        bg = alt_row_bg(i)
        tt = rel.get("tie_type", "")
        row_data = [
            rel.get("source", ""),
            rel.get("target", ""),
            tt.capitalize(),
            rel.get("sub_type", ""),
            rel.get("direction", ""),
            rel.get("strength", ""),
            rel.get("confidence", ""),
            ", ".join(str(x) for x in rel.get("chunk_indexes", [])),
            rel.get("evidence", "")[:200],
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col)
            body(cell, val, bg=bg, wrap=(col == 9))
        # Colour the tie-type cell
        tie_cell = ws.cell(row=i, column=3)
        color = TIE_COLORS.get(tt, "CCCCCC")
        tie_cell.fill = PatternFill("solid", start_color=color)
        tie_cell.font = Font(name="Arial", size=10, bold=True,
                             color=WHITE if tt in ("adversarial", "professional") else "000000")
        ws.row_dimensions[i].height = 40

    apply_border(ws, 1, 1 + len(rels), 1, len(headers))
    set_col_widths(ws, {
        "A": 20, "B": 20, "C": 14, "D": 18,
        "E": 12, "F": 11, "G": 12, "H": 10, "I": 50,
    })
    freeze(ws, "A2")
    ws.auto_filter.ref = f"A1:I{1 + len(rels)}"


def build_by_tie_type(wb: Workbook, data: dict):
    """One sheet per tie type showing only those relationships."""
    rels = data.get("relationships", [])
    for tie_type in ["family", "friendship", "romantic", "professional", "adversarial"]:
        subset = [r for r in rels if r.get("tie_type") == tie_type]
        if not subset:
            continue

        ws = wb.create_sheet(tie_type.capitalize())
        color = TIE_COLORS.get(tie_type, "CCCCCC")

        # Sheet title banner
        ws.merge_cells("A1:G1")
        hdr(ws["A1"], f"{tie_type.upper()} TIES  ({len(subset)})",
            bg=color, color=WHITE if tie_type in ("adversarial", "professional") else DARK_BLUE,
            size=12, center=True)
        ws.row_dimensions[1].height = 22

        headers = ["Source", "Target", "Sub-Type", "Direction",
                   "Strength", "Confidence", "Evidence"]
        for col, h in enumerate(headers, 1):
            sub_hdr(ws.cell(row=2, column=col), h)

        for i, rel in enumerate(subset, start=3):
            bg = alt_row_bg(i)
            row_data = [
                rel.get("source", ""),
                rel.get("target", ""),
                rel.get("sub_type", ""),
                rel.get("direction", ""),
                rel.get("strength", ""),
                rel.get("confidence", ""),
                rel.get("evidence", "")[:200],
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=i, column=col)
                body(cell, val, bg=bg, wrap=(col == 7))
            ws.row_dimensions[i].height = 40

        apply_border(ws, 2, 2 + len(subset), 1, len(headers))
        set_col_widths(ws, {
            "A": 20, "B": 20, "C": 18, "D": 12,
            "E": 11, "F": 12, "G": 52,
        })
        freeze(ws, "A3")
        ws.auto_filter.ref = f"A2:G{2 + len(subset)}"


def build_adjacency(wb: Workbook, data: dict):
    """One adjacency matrix sheet per tie type, plus an All-Ties overview."""
    characters = data.get("characters", [])
    rels = data.get("relationships", [])

    if not characters:
        return

    ids = [c.get("id", "") for c in characters]
    names = {c.get("id", ""): c.get("name", c.get("id", "")) for c in characters}

    def make_matrix(sheet_name: str, subset_rels: list[dict],
                    tie_type: str | None = None):
        """
        Build one matrix sheet.
        tie_type=None → all-ties overview (colour-coded by type).
        tie_type=str  → binary presence/absence for that one type.
        """
        # Always include ALL characters so isolates appear as all-zero rows/columns.
        # For per-type sheets, only filter if NO character at all has that tie type.
        display_ids = ids
        if not display_ids:
            return

        ws = wb.create_sheet(sheet_name)
        color = TIE_COLORS.get(tie_type, DARK_BLUE) if tie_type else DARK_BLUE
        font_color = WHITE if tie_type in ("adversarial", "professional") or tie_type is None else DARK_BLUE

        # Title banner — row 1
        n_cols = 1 + len(display_ids)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        banner_text = (f"{tie_type.upper()} ADJACENCY MATRIX" if tie_type
                       else "ALL-TIES ADJACENCY MATRIX")
        hdr(ws.cell(row=1, column=1), banner_text,
            bg=color, color=font_color, size=12, center=True)
        ws.row_dimensions[1].height = 22

        # Build lookup for this subset
        lookup: dict[tuple, list[str]] = {}
        for r in subset_rels:
            src = r.get("source", "")
            tgt = r.get("target", "")
            tt = r.get("tie_type", "")
            key = tuple(sorted([src, tgt]))
            lookup.setdefault(key, [])
            if tt not in lookup[key]:
                lookup[key].append(tt)

        isolate_ids = {c.get("id") for c in data.get("isolates", [])}

        # Column headers (rotated names)
        for col, cid in enumerate(display_ids, start=2):
            cell = ws.cell(row=2, column=col)
            is_isolate = cid in isolate_ids
            sub_hdr(cell, names.get(cid, cid))
            if is_isolate:
                cell.fill = PatternFill("solid", start_color="FFD7D7")
                cell.font = Font(name="Arial", size=10, bold=True, color="CC0000")
            cell.alignment = Alignment(horizontal="center", vertical="bottom",
                                       text_rotation=45, wrap_text=False)
        ws.row_dimensions[2].height = 80

        # Row headers + matrix cells
        for row, src_id in enumerate(display_ids, start=3):
            hdr_cell = ws.cell(row=row, column=1)
            is_isolate = src_id in isolate_ids
            hdr_bg = "FFD7D7" if is_isolate else LIGHT_BLUE
            body(hdr_cell, names.get(src_id, src_id), bold=True, bg=hdr_bg)
            if is_isolate:
                hdr_cell.font = Font(name="Arial", size=10, bold=True, color="CC0000")
            ws.row_dimensions[row].height = 18

            for col, tgt_id in enumerate(display_ids, start=2):
                cell = ws.cell(row=row, column=col)

                if src_id == tgt_id:
                    cell.fill = PatternFill("solid", start_color="D9D9D9")
                    cell.value = "—"
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.font = Font(name="Arial", size=10, color="888888")
                    continue

                key = tuple(sorted([src_id, tgt_id]))
                ties = lookup.get(key, [])
                bg = alt_row_bg(row)

                if tie_type:
                    # Binary: just show whether this specific tie exists
                    if tie_type in ties:
                        cell.fill = PatternFill("solid", start_color=color)
                        cell.value = 1
                        cell.font = Font(name="Arial", size=11, bold=True, color=font_color)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.value = 0
                        cell.font = Font(name="Arial", size=10, color="888888")
                        cell.fill = PatternFill("solid", start_color=bg)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    # All-ties: colour by type, multiplex gets yellow
                    if not ties:
                        body(cell, "", bg=bg)
                    elif len(ties) == 1:
                        tt = ties[0]
                        cell.fill = PatternFill("solid", start_color=TIE_COLORS.get(tt, "CCCCCC"))
                        cell.value = tt[:3].upper()
                        fc = WHITE if tt in ("adversarial", "professional") else DARK_BLUE
                        cell.font = Font(name="Arial", size=8, bold=True, color=fc)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        abbrevs = "+".join(t[:3].upper() for t in ties)
                        body(cell, abbrevs, bg="FFE699", center=True)
                        cell.font = Font(name="Arial", size=7, bold=True)

        apply_border(ws, 2, 2 + len(display_ids), 1, 1 + len(display_ids))

        ws.column_dimensions["A"].width = 22
        for col in range(2, 2 + len(display_ids)):
            ws.column_dimensions[get_column_letter(col)].width = 7

        # Legend (all-ties sheet only)
        if not tie_type:
            legend_row = 3 + len(display_ids) + 1
            ws.cell(row=legend_row, column=1).value = "Legend:"
            ws.cell(row=legend_row, column=1).font = Font(name="Arial", size=10, bold=True)
            for i, (tt, c) in enumerate(TIE_COLORS.items(), start=1):
                cell = ws.cell(row=legend_row, column=1 + i)
                cell.value = tt.capitalize()
                cell.fill = PatternFill("solid", start_color=c)
                cell.font = Font(name="Arial", size=9, bold=True,
                                 color=WHITE if tt in ("adversarial", "professional") else DARK_BLUE)
                cell.alignment = Alignment(horizontal="center")

    # All-ties overview
    make_matrix("All-Ties Matrix", rels, tie_type=None)

    # One sheet per tie type (only if that type has relationships)
    for tt in ["family", "friendship", "romantic", "professional", "adversarial"]:
        subset = [r for r in rels if r.get("tie_type") == tt]
        if subset:
            make_matrix(f"{tt.capitalize()} Matrix", subset, tie_type=tt)



def build_evidence(wb: Workbook, data: dict):
    """Full evidence detail — every span for every relationship."""
    ws = wb.create_sheet("Evidence Detail")
    rels = data.get("relationships", [])

    headers = ["Source", "Target", "Tie Type", "Sub-Type",
               "Chunk", "Text Snippet", "Reason"]
    for col, h in enumerate(headers, 1):
        sub_hdr(ws.cell(row=1, column=col), h)
    ws.row_dimensions[1].height = 18

    row = 2
    for rel in rels:
        spans = rel.get("evidence_spans", [])
        if not spans:
            spans = [{"chunk_index": "?", "text_snippet": rel.get("evidence", ""), "reason": ""}]
        for span in spans:
            bg = alt_row_bg(row)
            tt = rel.get("tie_type", "")
            row_data = [
                rel.get("source", ""),
                rel.get("target", ""),
                tt.capitalize(),
                rel.get("sub_type", ""),
                str(span.get("chunk_index", "")),
                span.get("text_snippet", "")[:300],
                span.get("reason", ""),
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col)
                body(cell, val, bg=bg, wrap=(col in (6, 7)))
            # Colour tie type cell
            tc = ws.cell(row=row, column=3)
            tc.fill = PatternFill("solid", start_color=TIE_COLORS.get(tt, "CCCCCC"))
            ws.row_dimensions[row].height = 45
            row += 1

    apply_border(ws, 1, row - 1, 1, len(headers))
    set_col_widths(ws, {
        "A": 20, "B": 20, "C": 14, "D": 18,
        "E": 8, "F": 55, "G": 35,
    })
    freeze(ws, "A2")
    ws.auto_filter.ref = f"A1:G{row - 1}"


def build_critique(wb: Workbook, data: dict):
    """Last critic round details."""
    critiques = data.get("critiques", [])
    if not critiques:
        return

    ws = wb.create_sheet("Critic Report")
    critique = critiques[-1]   # most recent round

    verdict = critique.get("quality_verdict", "?")
    verdict_colors = {"good": "70AD47", "needs_work": "FFC000", "poor": "FF0000"}
    vc = verdict_colors.get(verdict, "CCCCCC")

    ws.merge_cells("A1:E1")
    hdr(ws["A1"], f"Critic Report — Verdict: {verdict.upper()}",
        bg=vc, color=WHITE, size=13, center=True)
    ws.row_dimensions[1].height = 26

    def section(title, items, fields, row_start):
        ws.merge_cells(f"A{row_start}:E{row_start}")
        sub_hdr(ws[f"A{row_start}"], f"{title}  ({len(items)})")
        row_start += 1
        if not items:
            ws.merge_cells(f"A{row_start}:E{row_start}")
            body(ws[f"A{row_start}"], "None", bg=LIGHT_GREY)
            return row_start + 1
        for col, f in enumerate(fields, 1):
            hdr(ws.cell(row=row_start, column=col), f,
                bold=True, bg=LIGHT_BLUE, color=DARK_BLUE, size=9)
        row_start += 1
        for i, item in enumerate(items):
            bg = alt_row_bg(row_start)
            for col, f in enumerate(fields, 1):
                val = item.get(f, "")
                if isinstance(val, list):
                    val = ", ".join(str(x) for x in val)
                cell = ws.cell(row=row_start, column=col)
                body(cell, str(val), bg=bg, wrap=True)
                ws.row_dimensions[row_start].height = 35
            row_start += 1
        apply_border(ws, row_start - len(items) - 1, row_start - 1, 1, len(fields))
        return row_start + 1

    row = 3
    row = section("Gaps", critique.get("gaps", []),
                  ["source", "target", "suspected_tie_type", "reason", "recommended_chunk_indexes"], row)
    row = section("Low Confidence Flags", critique.get("low_confidence_flags", []),
                  ["source", "target", "tie_type", "concern", "recommended_chunk_indexes"], row)
    row = section("Contradictions", critique.get("contradictions", []),
                  ["description", "recommended_chunk_indexes"], row)
    row = section("Overreach Flags", critique.get("overreach_flags", []),
                  ["source", "target", "tie_type", "concern", "recommended_chunk_indexes"], row)
    row = section("ER Flags", critique.get("entity_resolution_flags", []),
                  ["from_id", "to_id", "concern", "recommended_chunk_indexes"], row)

    # Notes
    notes = critique.get("notes", [])
    if notes:
        ws.merge_cells(f"A{row}:E{row}")
        sub_hdr(ws[f"A{row}"], "Notes")
        row += 1
        for note in notes:
            ws.merge_cells(f"A{row}:E{row}")
            body(ws[f"A{row}"], f"• {note}", bg=alt_row_bg(row), wrap=True)
            ws.row_dimensions[row].height = 30
            row += 1

    set_col_widths(ws, {"A": 20, "B": 20, "C": 18, "D": 40, "E": 20})
    freeze(ws, "A2")


# ── Network export helpers ────────────────────────────────────────────────────

def sanitize_label(s: str) -> str:
    """UCINET-safe label: no spaces, no special chars."""
    s = s.strip().replace(" ", "_")
    s = re.sub(r"[^A-Za-z0-9_]", "", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s


def extract_edges_and_isolates(data: dict) -> tuple[list[tuple], list[str]]:
    """Pull (source, target, tie_type) edges and isolate names from JSON."""
    rels = data.get("relationships", [])
    edges = []
    for r in rels:
        src = r.get("source", "").strip()
        tgt = r.get("target", "").strip()
        tt  = r.get("tie_type", "").strip()
        if src and tgt and tt:
            edges.append((src, tgt, tt))

    isolates = [c.get("name", c.get("id", ""))
                for c in data.get("isolates", []) if c.get("id")]
    return edges, isolates


def export_ucinet(out_dir: Path, edges: list[tuple], isolates: list[str],
                  undirected: bool = True) -> Path:
    """Write a UCINET DL multiplex matrix file."""
    # Build node list
    nodes_set: set[str] = set()
    for u, v, _ in edges:
        nodes_set.add(u)
        nodes_set.add(v)
    for iso in isolates:
        nodes_set.add(iso)
    nodes = sorted(nodes_set)

    # Deduplicate edges
    seen: set[tuple] = set()
    clean: list[tuple] = []
    for u, v, lab in edges:
        a, b = (min(u, v), max(u, v)) if undirected else (u, v)
        key = (a, b, lab)
        if key not in seen:
            seen.add(key)
            clean.append((u, v, lab))

    # Collect ordered labels
    label_order: list[str] = []
    label_seen: set[str] = set()
    for _, _, lab in clean:
        if lab not in label_seen:
            label_order.append(lab)
            label_seen.add(lab)

    # Build matrices
    idx = {n: i for i, n in enumerate(nodes)}
    mats = {lab: [[0]*len(nodes) for _ in nodes] for lab in label_order}
    for u, v, lab in clean:
        i, j = idx[u], idx[v]
        mats[lab][i][j] = 1
        if undirected:
            mats[lab][j][i] = 1

    out_path = out_dir / "ucinet_multiplex.txt"
    nodes_s  = [sanitize_label(n) for n in nodes]
    labels_s = [sanitize_label(l) for l in label_order]

    with open(out_path, "w", encoding="utf-8") as f:
        f.write(f"dl n = {len(nodes_s)}, nm = {len(labels_s)}\n")
        f.write("labels:\n")
        f.write(",".join(nodes_s) + "\n")
        f.write("matrix labels:\n")
        f.write(",".join(labels_s) + "\n")
        f.write("data:\n")
        for k, lab in enumerate(label_order):
            for row in mats[lab]:
                f.write(" ".join(str(x) for x in row) + "\n")
            if k != len(label_order) - 1:
                f.write("\n")

    return out_path


def export_igraph(out_dir: Path, edges: list[tuple],
                  isolates: list[str], data: dict) -> tuple[Path, Path]:
    """Write iGraph-ready edges.csv and nodes.csv."""
    nodes_set: set[str] = set()
    for u, v, _ in edges:
        nodes_set.add(u)
        nodes_set.add(v)
    iso_set = set(isolates)
    nodes_set |= iso_set

    # Build id→name map from characters list
    id_to_name = {c.get("id", ""): c.get("name", c.get("id", ""))
                  for c in data.get("characters", [])}

    edges_path = out_dir / "edges_igraph.csv"
    nodes_path = out_dir / "nodes_igraph.csv"

    with open(edges_path, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["from", "to", "tie_type", "sub_type",
                    "direction", "strength", "confidence"])
        for r in data.get("relationships", []):
            w.writerow([
                r.get("source", ""),
                r.get("target", ""),
                r.get("tie_type", ""),
                r.get("sub_type", ""),
                r.get("direction", ""),
                r.get("strength", ""),
                r.get("confidence", ""),
            ])

    with open(nodes_path, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["id", "name", "role", "aliases", "is_isolate"])
        for c in data.get("characters", []):
            cid = c.get("id", "")
            w.writerow([
                cid,
                c.get("name", cid),
                c.get("role", ""),
                "|".join(c.get("aliases", [])),
                int(cid in {c2.get("id") for c2 in data.get("isolates", [])}),
            ])

    return edges_path, nodes_path


def export_networkx(out_dir: Path, edges: list[tuple],
                    data: dict) -> Path:
    """Write NetworkX GraphML, GEXF, CSVs, and per-layer GraphML files."""
    nx_dir = out_dir / "networkx"
    nx_dir.mkdir(exist_ok=True)

    # Build full node set with attributes
    chars = {c.get("id", ""): c for c in data.get("characters", [])}
    isolate_ids = {c.get("id") for c in data.get("isolates", [])}

    G = nx.MultiDiGraph()
    for cid, c in chars.items():
        G.add_node(cid,
                   name=c.get("name", cid),
                   role=c.get("role", ""),
                   is_isolate=int(cid in isolate_ids))

    for r in data.get("relationships", []):
        src = r.get("source", "")
        tgt = r.get("target", "")
        if src in G.nodes and tgt in G.nodes:
            G.add_edge(src, tgt,
                       tie_type=r.get("tie_type", ""),
                       sub_type=r.get("sub_type", ""),
                       direction=r.get("direction", ""),
                       strength=r.get("strength", ""),
                       confidence=r.get("confidence", ""),
                       evidence=r.get("evidence", "")[:200])

    # Full multiplex exports
    nx.write_graphml(G, nx_dir / "multiplex.graphml")
    nx.write_gexf(G,   nx_dir / "multiplex.gexf")

    # Edge CSV
    with open(nx_dir / "edges_nx.csv", "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["source", "target", "tie_type", "sub_type",
                    "direction", "strength", "confidence"])
        for u, v, d in G.edges(data=True):
            w.writerow([u, v, d.get("tie_type",""), d.get("sub_type",""),
                        d.get("direction",""), d.get("strength",""),
                        d.get("confidence","")])

    # Node CSV
    with open(nx_dir / "nodes_nx.csv", "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["id", "name", "role", "is_isolate"])
        for n, d in G.nodes(data=True):
            w.writerow([n, d.get("name",""), d.get("role",""),
                        d.get("is_isolate", 0)])

    # Per-layer GraphML
    layers_dir = nx_dir / "layers"
    layers_dir.mkdir(exist_ok=True)
    layers: dict[str, list] = defaultdict(list)
    for u, v, d in G.edges(data=True):
        layers[d.get("tie_type", "unknown")].append((u, v, d))

    for tt, layer_edges in layers.items():
        Gl = nx.MultiDiGraph()
        for n, nd in G.nodes(data=True):
            Gl.add_node(n, **nd)
        for u, v, d in layer_edges:
            Gl.add_edge(u, v, **d)
        safe = re.sub(r"[^A-Za-z0-9_-]", "_", tt)
        nx.write_graphml(Gl, layers_dir / f"{safe}.graphml")

    return nx_dir


# ── Orchestrator ──────────────────────────────────────────────────────────────

def run_pipeline(json_path: str, out_root: str | None = None):
    data = json.loads(Path(json_path).read_text(encoding="utf-8"))

    story_name = data.get("story_name") or Path(json_path).stem
    data["story_name"] = story_name

    # Output folder: outputs/<story_stem>/
    if out_root:
        out_dir = Path(out_root) / story_name
    else:
        out_dir = Path(json_path).parent / story_name
    out_dir.mkdir(parents=True, exist_ok=True)

    edges, isolates = extract_edges_and_isolates(data)

    # ── Excel report ──
    wb = Workbook()
    wb.remove(wb.active)
    build_summary(wb, data)
    build_characters(wb, data)
    build_relationships(wb, data)
    build_by_tie_type(wb, data)
    build_adjacency(wb, data)
    build_evidence(wb, data)
    build_critique(wb, data)
    xl_path = out_dir / f"{story_name}.xlsx"
    wb.save(xl_path)
    print(f"  Excel report : {xl_path}")

    # ── UCINET ──
    uc_path = export_ucinet(out_dir, edges, isolates)
    print(f"  UCINET DL    : {uc_path}")

    # ── iGraph ──
    ig_edges, ig_nodes = export_igraph(out_dir, edges, isolates, data)
    print(f"  iGraph CSVs  : {ig_edges}, {ig_nodes}")

    # ── NetworkX ──
    nx_dir = export_networkx(out_dir, edges, data)
    print(f"  NetworkX     : {nx_dir}")

    print(f"\nAll outputs in: {out_dir.resolve()}")
    return out_dir


def main():
    parser = argparse.ArgumentParser(
        description="Convert SNA JSON → Excel + UCINET + iGraph + NetworkX"
    )
    parser.add_argument("json", help="Path to network JSON file")
    parser.add_argument("--outdir", default=None,
                        help="Root output directory (default: same folder as JSON)")
    args = parser.parse_args()

    if not Path(args.json).exists():
        print(f"Error: file not found: {args.json}", file=sys.stderr)
        sys.exit(1)

    print(f"Processing: {args.json}")
    run_pipeline(args.json, args.outdir)


if __name__ == "__main__":
    main()
